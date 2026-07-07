"""Flask Web应用 - 路由和视图"""
import json
import time
import uuid
import logging
import os
import shutil
import tempfile
import threading
import queue
from datetime import datetime
from flask import Blueprint, request, jsonify, render_template, Response, stream_with_context
import sys
import io

# 确保UTF-8编码（PyInstaller无控制台模式下stdout/stderr可能为None）
if sys.stdout and hasattr(sys.stdout, 'buffer') and sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
if sys.stderr and hasattr(sys.stderr, 'buffer') and sys.stderr.encoding != 'utf-8':
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# 项目根目录（青稞/）
if getattr(sys, 'frozen', False):
    # PyInstaller 单文件模式：可写数据放在exe所在目录
    _PROJECT_ROOT = os.path.dirname(sys.executable)
else:
    _PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
_UPLOAD_DIR = os.path.join(_PROJECT_ROOT, 'data', 'uploads')
_LOG_DIR = os.path.join(_PROJECT_ROOT, 'data', 'logs')

# 确保日志目录存在
os.makedirs(_LOG_DIR, exist_ok=True)

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(_LOG_DIR, 'api.log'), encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger('api')


def _append_task_event(entry, event):
    """向任务事件列表追加事件（线程安全）"""
    with entry['lock']:
        entry['events'].append(event)


def _run_task_background(user_input, session_id, storage_entry, engine_ref, abort_set):
    """后台线程：执行对话任务并将事件写入共享存储（轮询模式替代SSE长连接）"""
    try:
        abort_set.discard(session_id)
        engine_ref.agentic_loop._abort_requested = False

        # 1. 前台接待分析
        agent_result = engine_ref.front_desk.process(user_input, session_id)
        _append_task_event(storage_entry, {
            'type': 'analysis',
            'action': agent_result['action'],
            'dag_suggested': agent_result.get('dag_suggested', False)
        })

        if agent_result["action"] == "direct_reply":
            response = agent_result.get("reply", "")
            engine_ref.pref_learner.record_turn(user_input, response)
            engine_ref.context_mgr.save_conversation(session_id, user_input, response)
            _append_task_event(storage_entry, {
                'type': 'response', 'response': response,
                'session_id': session_id, 'need_dag': False
            })

        elif agent_result["action"] == "need_info":
            missing = agent_result.get("missing_info", [])
            response = f"需要补充以下信息：{', '.join(missing)}"
            engine_ref.context_mgr.save_conversation(session_id, user_input, response)
            engine_ref.pref_learner.record_turn(user_input, response)
            _append_task_event(storage_entry, {
                'type': 'response', 'response': response,
                'session_id': session_id, 'need_dag': False
            })

        elif agent_result["action"] == "need_dag":
            use_agentic = engine_ref.config.get("agentic_loop", {}).get("enabled", True)

            if use_agentic:
                final_response = ""
                steps = []
                today = datetime.now().strftime("%Y-%m-%d")
                engine_ref.logger.log_session_start(session_id, user_input, today)
                engine_ref.logger.log_agent_analysis(agent_result["action"], agent_result, today)

                for evt in engine_ref.agentic_loop.run(user_input, session_id):
                    if session_id in abort_set:
                        break
                    _append_task_event(storage_entry, evt)
                    steps.append(evt)

                    if evt.get("type") == "dag_node_complete":
                        if evt.get("name") == "回复用户" or evt.get("command") == "reply_to_user":
                            final_response = evt.get("result", "")
                        elif evt.get("name") == "任务完成" or evt.get("command") == "task_complete":
                            if not final_response:
                                final_response = evt.get("result", "")
                        engine_ref._incremental_save(session_id, user_input, steps, evt, today)

                # 处理终止
                if session_id in abort_set:
                    abort_set.discard(session_id)
                    abort_summary = "（用户终止了DAG执行）\n\n"
                    if steps:
                        completed = [s for s in steps if s.get("type") == "dag_node_complete"]
                        if completed:
                            abort_summary += "已执行的步骤结果：\n"
                            for s in completed:
                                abort_summary += f"- {s.get('name', '')}: {s.get('result', '')[:100]}\n"
                    final_response = final_response or abort_summary
                    _append_task_event(storage_entry, {
                        'type': 'aborted', 'response': final_response,
                        'session_id': session_id
                    })
                    storage_entry['status'] = 'aborted'
                    return

                # 保存对话
                try:
                    engine_ref.context_mgr.save_conversation(session_id, user_input, final_response)
                    engine_ref.pref_learner.record_turn(user_input, final_response)
                    engine_ref._save_agentic_as_dag(session_id, user_input, steps)
                except Exception as save_err:
                    logger.error(f"保存对话/DAG记录失败: {save_err}", exc_info=True)

                _append_task_event(storage_entry, {
                    'type': 'done', 'response': final_response,
                    'session_id': session_id, 'need_dag': True, 'mode': 'agentic_loop'
                })

            else:
                # 传统 DAG 模式
                result = engine_ref._handle_dag_flow(user_input, session_id, agent_result, datetime.now().strftime("%Y-%m-%d"))
                _append_task_event(storage_entry, {
                    'type': 'response', 'response': result.get('response', ''),
                    'session_id': session_id, 'need_dag': True, 'dag_id': result.get('dag_id')
                })

        storage_entry['status'] = 'done'

    except Exception as e:
        logger.error(f"后台任务异常: {e}", exc_info=True)
        _append_task_event(storage_entry, {
            'type': 'error', 'message': f'服务内部异常: {str(e)[:200]}'
        })
        storage_entry['status'] = 'error'


def create_api_blueprint(engine):
    """创建API蓝图"""
    api = Blueprint('api', __name__, url_prefix='/api')

    # 终止标志集合：记录请求终止的session_id
    _abort_sessions = set()
    # 轮询模式任务存储：{session_id: {'status': str, 'events': list, 'lock': Lock, 'started_at': float}}
    _task_storage = {}

    @api.route('/chat', methods=['POST'])
    def chat():
        """发送对话消息"""
        # 确保正确处理UTF-8编码的JSON
        raw_data = request.get_data(as_text=True)
        try:
            data = json.loads(raw_data)
        except json.JSONDecodeError:
            data = request.json or {}

        user_input = data.get("input", "")
        session_id = data.get("session_id", f"session_{uuid.uuid4().hex[:8]}")
        files = data.get("files", [])  # 接收文件列表（含file_text）

        if not user_input and not files:
            return jsonify({"error": "输入不能为空"}), 400

        # 如果有文件，将文件解析文本附加到用户输入
        if files:
            file_parts = []
            for f in files:
                if f.get('file_text'):
                    file_parts.append(f['file_text'])
                else:
                    file_parts.append(f"「这里有个文件，路径为：{f.get('path', '')}，文件名为：{f.get('name', '')}」")
            file_info = "\n\n" + "\n".join(file_parts)
            user_input = user_input + file_info if user_input else file_info

        result = engine.process(user_input, session_id)
        return jsonify({
            "response": result.get("response", ""),
            "session_id": session_id,
            "need_dag": result.get("need_dag", False),
            "dag_id": result.get("dag_id"),
        })

    @api.route('/chat/abort', methods=['POST'])
    def chat_abort():
        """终止当前DAG执行"""
        raw_data = request.get_data(as_text=True)
        try:
            data = json.loads(raw_data)
        except json.JSONDecodeError:
            data = request.json or {}

        session_id = data.get("session_id", "")
        if session_id:
            _abort_sessions.add(session_id)
            # 设置 AgenticLoop 的终止标志
            engine.agentic_loop._abort_requested = True
            logger.info(f"收到终止请求: session={session_id}")
            return jsonify({"status": "abort_requested", "session_id": session_id})
        return jsonify({"error": "缺少session_id"}), 400

    @api.route('/chat/respond', methods=['POST'])
    def chat_respond():
        """用户回复ask_user的提问（DAG中间节点等待用户输入时调用）"""
        raw_data = request.get_data(as_text=True)
        try:
            data = json.loads(raw_data)
        except json.JSONDecodeError:
            data = request.json or {}

        answer = data.get("answer", "").strip()
        session_id = data.get("session_id", "")

        if not answer:
            return jsonify({"error": "回复内容不能为空"}), 400

        # 将用户回复传递给 AgenticLoop
        if engine.agentic_loop.is_waiting_for_user:
            engine.agentic_loop.provide_user_response(answer)
            logger.info(f"用户回复ask_user: session={session_id}, answer={answer[:100]}")
            return jsonify({"status": "ok", "session_id": session_id})
        else:
            return jsonify({"error": "当前没有等待用户回复的任务"}), 400

    @api.route('/select-folder', methods=['POST'])
    def select_folder():
        """验证工作区文件夹路径是否存在"""
        data = request.get_json(silent=True) or {}
        folder_path = data.get("path", "").strip()
        if folder_path and os.path.isdir(folder_path):
            return jsonify({"status": "ok", "path": os.path.abspath(folder_path)})
        elif folder_path:
            return jsonify({"status": "error", "message": f"路径不存在: {folder_path}"})
        else:
            return jsonify({"status": "cancelled", "path": ""})

    @api.route('/chat/start', methods=['POST'])
    def chat_start():
        """启动异步对话任务（轮询模式，替代SSE长连接，彻底解决Network Error）"""
        raw_data = request.get_data(as_text=True)
        try:
            data = json.loads(raw_data)
        except json.JSONDecodeError:
            data = request.json or {}

        user_input = data.get("input", "")
        session_id = data.get("session_id", f"session_{uuid.uuid4().hex[:8]}")
        files = data.get("files", [])

        if not user_input and not files:
            return jsonify({"error": "输入不能为空"}), 400

        # 文件处理
        if files:
            file_parts = []
            for f in files:
                if f.get('file_text'):
                    file_parts.append(f['file_text'])
                else:
                    file_parts.append(f"「这里有个文件，路径为：{f.get('path', '')}，文件名为：{f.get('name', '')}」")
            file_info = "\n\n" + "\n".join(file_parts)
            user_input = user_input + file_info if user_input else file_info

        # 如果该session已有运行中的任务，拒绝
        existing = _task_storage.get(session_id)
        if existing and existing['status'] == 'running':
            return jsonify({"error": "该会话已有运行中的任务"}), 409

        # 创建任务存储
        _task_storage[session_id] = {
            'status': 'running',
            'events': [],
            'lock': threading.Lock(),
            'started_at': time.time()
        }

        # 启动后台线程执行任务
        storage_entry = _task_storage[session_id]
        t = threading.Thread(
            target=_run_task_background,
            args=(user_input, session_id, storage_entry, engine, _abort_sessions),
            daemon=True
        )
        t.start()

        logger.info(f"轮询任务已启动: session={session_id}")
        return jsonify({"status": "started", "session_id": session_id})

    @api.route('/chat/poll/<session_id>', methods=['GET'])
    def chat_poll(session_id):
        """轮询获取任务事件（短轮询替代SSE长连接，彻底解决Network Error）"""
        since = request.args.get('since', 0, type=int)
        task = _task_storage.get(session_id)
        if not task:
            return jsonify({"status": "not_found", "events": [], "total": 0})

        with task['lock']:
            events = list(task['events'][since:])
            status = task['status']
            total = len(task['events'])

        return jsonify({"status": status, "events": events, "total": total})

    @api.route('/chat/cleanup/<session_id>', methods=['DELETE'])
    def chat_cleanup(session_id):
        """清理已完成的任务数据"""
        _task_storage.pop(session_id, None)
        return jsonify({"status": "cleaned"})

    @api.route('/dag/list', methods=['GET'])
    def dag_list():
        """获取DAG列表"""
        dags = engine.dag_parser.list_dags()
        return jsonify(dags)

    @api.route('/dag/<dag_id>', methods=['GET'])
    def dag_detail(dag_id):
        """获取DAG详情"""
        dag = engine.dag_parser.load(dag_id)
        if not dag:
            return jsonify({"error": "DAG not found"}), 404
        return jsonify({
            "id": dag.id,
            "name": dag.name,
            "description": dag.description,
            "status": dag.status,
            "nodes": [
                {
                    "id": n.id,
                    "name": n.name,
                    "status": n.status,
                    "dependencies": n.dependencies,
                    "retry_count": n.retry_count,
                    "result": n.result,
                }
                for n in dag.nodes.values()
            ]
        })

    @api.route('/dag/<dag_id>/status', methods=['GET'])
    def dag_status(dag_id):
        """获取DAG执行状态（SSE推送）"""
        def generate():
            while True:
                status = engine.scheduler.get_dag_status(dag_id)
                event_data = json.dumps(status, ensure_ascii=False)
                yield f"data: {event_data}\n\n"

                if status.get("status") in ("completed", "failed"):
                    break
                time.sleep(1)

        return Response(stream_with_context(generate()), mimetype='text/event-stream')

    @api.route('/dag/<dag_id>/run', methods=['POST'])
    def dag_run(dag_id):
        """执行DAG"""
        try:
            results = []
            for node in engine.scheduler.run_dag(dag_id):
                results.append({
                    "id": node.id,
                    "name": node.name,
                    "status": node.status,
                    "result": node.result,
                })
            return jsonify({"results": results})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @api.route('/dag/<dag_id>/retry', methods=['POST'])
    def dag_retry(dag_id):
        """重试DAG（可选指定节点）"""
        data = request.json or {}
        node_id = data.get("node_id")

        try:
            results = []
            for node in engine.scheduler.retry_dag_node(dag_id, node_id):
                results.append({
                    "id": node.id,
                    "name": node.name,
                    "status": node.status,
                    "result": node.result,
                })
            return jsonify({"results": results})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @api.route('/logs', methods=['GET'])
    def logs_list():
        """获取日志列表"""
        logs = engine.logger.get_logs()
        return jsonify(logs)

    @api.route('/logs/<date>', methods=['GET'])
    def logs_detail(date):
        """获取指定日期日志"""
        content = engine.logger.get_log_content(date)
        if not content:
            return jsonify({"error": "Log not found"}), 404
        return jsonify({"content": content})

    @api.route('/preferences', methods=['GET'])
    def preferences_get():
        """获取当前用户偏好"""
        prefs = engine.pref_learner.get_preferences()
        return jsonify(prefs)

    @api.route('/preferences/review', methods=['POST'])
    def preferences_review():
        """审核偏好更新（确认/拒绝/编辑）"""
        data = request.json
        action = data.get("action")  # confirm | reject | edit
        updates = data.get("updates", [])

        if action == "confirm":
            engine.pref_learner._update_preferences(updates)
            engine.pref_learner._update_sys_prompt(updates)
            engine.pref_learner._reload_sys_prompt()
            return jsonify({"status": "updated"})
        elif action == "reject":
            return jsonify({"status": "rejected"})
        elif action == "edit":
            # 用户编辑后的更新
            engine.pref_learner._update_preferences(updates)
            engine.pref_learner._update_sys_prompt(updates)
            return jsonify({"status": "updated"})
        else:
            return jsonify({"error": "Invalid action"}), 400

    @api.route('/preferences/rollback', methods=['POST'])
    def preferences_rollback():
        """回滚偏好到指定版本"""
        data = request.json
        timestamp = data.get("timestamp", "")
        success = engine.pref_learner.rollback(timestamp)
        if success:
            return jsonify({"status": "rolled back"})
        return jsonify({"error": "Rollback failed"}), 404

    @api.route('/preferences/history', methods=['GET'])
    def preferences_history():
        """获取偏好变更历史"""
        history = engine.pref_learner.get_history()
        return jsonify(history)

    @api.route('/config', methods=['GET'])
    def config_get():
        """获取当前配置"""
        return jsonify(engine.config)

    @api.route('/config', methods=['POST'])
    def config_update():
        """更新配置（带校验和自动备份）"""
        try:
            data = request.json or {}

            # 1. 校验并修正前端提交的配置
            cleaned = engine.config_guard.validate_partial_update(data)
            if not cleaned:
                return jsonify({"error": "配置项无效或为空"}), 400

            # 2. 合并到内存配置
            engine.config.update(cleaned)

            # 3. 动态更新 LLM 网关配置
            if "llm" in cleaned:
                engine.llm.update_config(engine.config.get("llm", {}))

            # 4. 持久化到文件（带自动备份）
            engine.config_guard.save_config_with_backup(engine.config)

            logger.info(f"配置更新成功 - {json.dumps(cleaned, ensure_ascii=False)[:200]}")
            return jsonify({"status": "updated"})
        except Exception as e:
            logger.error(f"配置更新失败: {str(e)}")
            return jsonify({"error": f"保存失败：{str(e)}"}), 500

    @api.route('/config/reset', methods=['POST'])
    def config_reset():
        """恢复默认配置（先备份再重置）"""
        try:
            # 备份当前配置再重置
            engine.config_guard.save_config_with_backup(engine.config_guard.load_and_validate_config())
            from core.config_guard import _DEFAULT_CONFIG
            from copy import deepcopy
            default_config = deepcopy(_DEFAULT_CONFIG)

            # 更新内存和文件
            engine.config.update(default_config)
            engine.config_guard.save_config_with_backup(default_config)
            engine.llm.update_config(default_config.get("llm", {}))

            return jsonify(default_config)
        except Exception as e:
            logger.error(f"重置配置失败: {e}")
            return jsonify({"error": str(e)}), 500

    @api.route('/config/system', methods=['GET'])
    def config_system_info():
        """获取系统信息"""
        import sys
        import flask
        
        return jsonify({
            "appVersion": "1.0.0",
            "pythonVersion": f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}",
            "flaskVersion": flask.__version__
        })

    @api.route('/file/verify', methods=['POST'])
    def verify_file():
        """验证文件是否存在并返回文件信息"""
        data = request.json or {}
        file_path = data.get("path", "")
        
        if not file_path:
            return jsonify({"error": "文件路径不能为空"}), 400
        
        # 规范化路径
        file_path = os.path.normpath(file_path)
        
        if not os.path.exists(file_path):
            return jsonify({"error": "文件不存在", "path": file_path}), 404
        
        if not os.path.isfile(file_path):
            return jsonify({"error": "路径不是文件", "path": file_path}), 400
        
        try:
            stat = os.stat(file_path)
            file_name = os.path.basename(file_path)
            file_ext = os.path.splitext(file_name)[1].lower()
            
            return jsonify({
                "success": True,
                "path": file_path,
                "name": file_name,
                "size": stat.st_size,
                "extension": file_ext,
                "modified": stat.st_mtime
            })
        except Exception as e:
            return jsonify({"error": f"获取文件信息失败: {str(e)}"}), 500

    def _parse_file_content(file_path: str, file_name: str, file_ext: str) -> str:
        """解析文件内容，返回格式化的文本"""
        max_size = 50 * 1024  # 最大读取50KB

        # 二进制文件类型（不尝试读取内容）
        binary_exts = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.ico', '.svg',
                       '.mp3', '.wav', '.mp4', '.avi', '.mov', '.wmv',
                       '.zip', '.rar', '.7z', '.tar', '.gz',
                       '.exe', '.dll', '.so', '.bin', '.dat'}

        if file_ext in binary_exts:
            return f"（二进制文件，类型：{file_ext}，大小：{os.path.getsize(file_path)}字节，无法解析为文本）"

        # Excel 文件特殊处理
        if file_ext in ('.xlsx', '.xls'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                result_lines = []
                for sheet_name in wb.sheetnames[:3]:  # 最多读3个sheet
                    ws = wb[sheet_name]
                    result_lines.append(f"[Sheet: {sheet_name}]")
                    row_count = 0
                    for row in ws.iter_rows(max_row=50, values_only=True):  # 最多50行
                        row_text = "\t".join([str(c) if c is not None else "" for c in row])
                        result_lines.append(row_text)
                        row_count += 1
                    if row_count >= 50:
                        result_lines.append(f"（... 共{ws.max_row}行，仅显示前50行）")
                    result_lines.append("")
                wb.close()
                return "\n".join(result_lines)
            except ImportError:
                return "（Excel文件，需要openpyxl库才能解析）"
            except Exception as e:
                return f"（Excel解析失败：{str(e)}）"

        # CSV 文件
        if file_ext == '.csv':
            try:
                import csv
                result_lines = []
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    for i, row in enumerate(reader):
                        if i >= 100:  # 最多100行
                            result_lines.append("（... 仅显示前100行）")
                            break
                        result_lines.append("\t".join(row))
                return "\n".join(result_lines)
            except UnicodeDecodeError:
                # 尝试GBK编码
                try:
                    result_lines = []
                    with open(file_path, 'r', encoding='gbk') as f:
                        reader = csv.reader(f)
                        for i, row in enumerate(reader):
                            if i >= 100:
                                result_lines.append("（... 仅显示前100行）")
                                break
                            result_lines.append("\t".join(row))
                    return "\n".join(result_lines)
                except Exception as e:
                    return f"（CSV解析失败：{str(e)}）"

        # 文本文件（代码、配置、文档等）
        text_exts = {'.txt', '.md', '.py', '.js', '.ts', '.jsx', '.tsx',
                     '.html', '.css', '.json', '.yaml', '.yml', '.xml',
                     '.toml', '.ini', '.cfg', '.conf',
                     '.sh', '.bat', '.cmd', '.ps1',
                     '.c', '.cpp', '.h', '.hpp', '.java', '.go', '.rs',
                     '.sql', '.r', '.rb', '.php', '.swift', '.kt',
                     '.log', '.env', '.gitignore', '.dockerfile'}

        if file_ext in text_exts or file_ext == '' or file_ext == '.':
            try:
                file_size = os.path.getsize(file_path)
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read(max_size)
                if file_size > max_size:
                    content += f"\n\n（... 文件过大，仅显示前{max_size // 1024}KB，总大小{file_size // 1024}KB）"
                return content
            except UnicodeDecodeError:
                try:
                    with open(file_path, 'r', encoding='gbk') as f:
                        content = f.read(max_size)
                    return content
                except Exception:
                    return f"（文件编码不支持，无法解析为文本）"
            except Exception as e:
                return f"（文件读取失败：{str(e)}）"

        # 其他文件类型，尝试作为文本读取
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read(max_size)
            return content
        except (UnicodeDecodeError, Exception):
            return f"（未知文件类型：{file_ext}，无法解析为文本）"

    def _build_file_text(file_path: str, file_name: str, file_ext: str, content: str) -> str:
        """构建文件引用文本，使用中文书名号避免与代码解释符号冲突"""
        # 使用「」作为文件引用定界符，避免与反引号 ` 和引号 " 冲突
        return f"「这里有个文件，路径为：{file_path}，文件内容如下：\n{content}」"

    @api.route('/file/upload', methods=['POST'])
    def upload_file():
        """上传文件并解析内容"""
        os.makedirs(_UPLOAD_DIR, exist_ok=True)

        if 'file' not in request.files:
            return jsonify({"error": "没有文件"}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "文件名为空"}), 400

        try:
            # 保存文件到上传目录
            file_name = file.filename
            # 生成唯一文件名避免冲突
            unique_name = f"{uuid.uuid4().hex[:8]}_{file_name}"
            save_path = os.path.join(_UPLOAD_DIR, unique_name)
            file.save(save_path)

            file_ext = os.path.splitext(file_name)[1].lower()
            file_size = os.path.getsize(save_path)

            # 解析文件内容
            content = _parse_file_content(save_path, file_name, file_ext)

            # 构建文件引用文本
            file_text = _build_file_text(save_path, file_name, file_ext, content)

            return jsonify({
                "success": True,
                "name": file_name,
                "path": save_path,
                "size": file_size,
                "extension": file_ext,
                "content_preview": content[:500] + ("..." if len(content) > 500 else ""),
                "file_text": file_text
            })
        except Exception as e:
            logger.error(f"文件上传失败: {str(e)}")
            return jsonify({"error": f"文件上传失败: {str(e)}"}), 500

    @api.route('/file/upload/path', methods=['POST'])
    def upload_file_by_path():
        """通过本地路径解析文件内容（不上传，直接读取本地文件）"""
        data = request.json or {}
        file_path = data.get("path", "")

        if not file_path:
            return jsonify({"error": "文件路径不能为空"}), 400

        file_path = os.path.normpath(file_path)

        if not os.path.exists(file_path):
            return jsonify({"error": "文件不存在", "path": file_path}), 404

        if not os.path.isfile(file_path):
            return jsonify({"error": "路径不是文件", "path": file_path}), 400

        try:
            file_name = os.path.basename(file_path)
            file_ext = os.path.splitext(file_name)[1].lower()
            file_size = os.path.getsize(file_path)

            # 解析文件内容
            content = _parse_file_content(file_path, file_name, file_ext)

            # 构建文件引用文本
            file_text = _build_file_text(file_path, file_name, file_ext, content)

            return jsonify({
                "success": True,
                "name": file_name,
                "path": file_path,
                "size": file_size,
                "extension": file_ext,
                "content_preview": content[:500] + ("..." if len(content) > 500 else ""),
                "file_text": file_text
            })
        except Exception as e:
            logger.error(f"文件解析失败: {str(e)}")
            return jsonify({"error": f"文件解析失败: {str(e)}"}), 500

    @api.route('/file/remove', methods=['POST'])
    def remove_uploaded_file():
        """移除已上传的文件"""
        data = request.json or {}
        file_path = data.get("path", "")

        if not file_path:
            return jsonify({"error": "文件路径不能为空"}), 400

        # 安全检查：只允许删除上传目录中的文件
        file_path = os.path.normpath(file_path)
        if not file_path.startswith(os.path.normpath(_UPLOAD_DIR)):
            return jsonify({"error": "只能删除上传目录中的文件"}), 403

        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                return jsonify({"success": True, "message": "文件已删除"})
            else:
                return jsonify({"success": True, "message": "文件不存在，可能已被删除"})
        except Exception as e:
            return jsonify({"error": f"删除失败: {str(e)}"}), 500

    # ============================================================
    # Prompt 管理接口
    # ============================================================

    @api.route('/prompt', methods=['GET'])
    def prompt_get():
        """获取当前系统提示词内容"""
        try:
            prompt_path = os.path.join(_PROJECT_ROOT, 'config', 'sys_prompt.md')
            if not os.path.exists(prompt_path):
                return jsonify({"content": "", "exists": False})
            with open(prompt_path, 'r', encoding='utf-8') as f:
                content = f.read()
            return jsonify({"content": content, "exists": True})
        except Exception as e:
            return jsonify({"error": f"读取失败: {str(e)}"}), 500

    @api.route('/prompt', methods=['POST'])
    def prompt_save():
        """保存用户编辑的系统提示词（自动备份旧版本）"""
        try:
            data = request.json or {}
            content = data.get("content", "")
            if not content.strip():
                return jsonify({"error": "提示词内容不能为空"}), 400

            prompt_path = os.path.join(_PROJECT_ROOT, 'config', 'sys_prompt.md')
            backup_dir = os.path.join(_PROJECT_ROOT, 'config', 'backup')
            os.makedirs(backup_dir, exist_ok=True)

            # 备份当前文件
            if os.path.exists(prompt_path):
                ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                backup_path = os.path.join(backup_dir, f'sys_prompt_{ts}.md')
                shutil.copy2(prompt_path, backup_path)

            # 写入新内容
            os.makedirs(os.path.dirname(prompt_path), exist_ok=True)
            with open(prompt_path, 'w', encoding='utf-8') as f:
                f.write(content)

            # 热重载：更新引擎中的系统提示词
            engine.context_mgr.system_prompt = content

            logger.info("系统提示词已更新")
            return jsonify({"status": "saved"})
        except Exception as e:
            logger.error(f"保存提示词失败: {e}")
            return jsonify({"error": f"保存失败: {str(e)}"}), 500

    @api.route('/prompt/default', methods=['GET'])
    def prompt_default():
        """获取初始化默认模板"""
        try:
            from core.config_guard import _DEFAULT_PROMPT_TEMPLATE
            return jsonify({"content": _DEFAULT_PROMPT_TEMPLATE})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @api.route('/prompt/reset', methods=['POST'])
    def prompt_reset():
        """恢复为初始化默认模板（备份当前版本）"""
        try:
            from core.config_guard import _DEFAULT_PROMPT_TEMPLATE
            prompt_path = os.path.join(_PROJECT_ROOT, 'config', 'sys_prompt.md')
            backup_dir = os.path.join(_PROJECT_ROOT, 'config', 'backup')
            os.makedirs(backup_dir, exist_ok=True)

            # 备份当前文件
            if os.path.exists(prompt_path):
                ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                backup_path = os.path.join(backup_dir, f'sys_prompt_{ts}.md')
                shutil.copy2(prompt_path, backup_path)

            # 写入默认模板
            with open(prompt_path, 'w', encoding='utf-8') as f:
                f.write(_DEFAULT_PROMPT_TEMPLATE)

            # 热重载
            engine.context_mgr.system_prompt = _DEFAULT_PROMPT_TEMPLATE

            logger.info("系统提示词已重置为默认模板")
            return jsonify({"status": "reset", "content": _DEFAULT_PROMPT_TEMPLATE})
        except Exception as e:
            logger.error(f"重置提示词失败: {e}")
            return jsonify({"error": f"重置失败: {str(e)}"}), 500

    @api.route('/prompt/backups', methods=['GET'])
    def prompt_backups():
        """获取提示词备份列表"""
        try:
            backup_dir = os.path.join(_PROJECT_ROOT, 'config', 'backup')
            if not os.path.exists(backup_dir):
                return jsonify({"backups": []})
            backups = sorted(
                [f for f in os.listdir(backup_dir) if f.startswith('sys_prompt_') and f.endswith('.md')],
                reverse=True
            )
            return jsonify({"backups": backups})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    return api
